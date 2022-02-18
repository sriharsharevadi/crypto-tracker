from datetime import datetime

import openpyxl
from multiprocessing import Pool

import pytz

from crypto.wazirx_api_helpers import get_current_balance, get_price_at_time
from webapp.models import Coin


def get_list_from_excel(path, sheet_name, user_id):
    ist = pytz.timezone('Asia/Kolkata')

    results = []
    wb_obj = openpyxl.load_workbook(path)
    ech_trades = wb_obj.get_sheet_by_name(sheet_name)

    for row in ech_trades.iter_rows(min_row=2, min_col=1, max_row=ech_trades.max_row, max_col=ech_trades.max_column):
        temp = dict()
        temp["time"] = datetime.strptime(row[0].value, "%Y-%m-%d %H:%M:%S").replace(tzinfo=pytz.timezone('UTC')).astimezone(ist)
        temp["market"] = row[1].value
        temp["price"] = row[2].value
        temp["volume"] = row[3].value
        temp["total"] = row[4].value
        temp["type"] = row[5].value
        temp["fee_currency"] = row[6].value
        temp["fee"] = row[7].value
        temp["user"] = user_id
        results.append(temp)

    return results

# path = "harsha.xlsx"

# wb_obj = openpyxl.load_workbook(path)

# ech_trades = wb_obj.get_sheet_by_name('Exchange Trades')
# acc_balance = wb_obj.get_sheet_by_name('Account Balance')

# iterate through excel and display data
# coin_balance = {}
# for row in acc_balance.iter_rows(min_row=2, min_col=1, max_row=acc_balance.max_row, max_col=2):
#     coin_balance[row[0].value] = row[1].value

# result = get_current_balance(coin_balance)


def get_network_fee(trade):
    if "wrx" in trade["fee_currency"].lower():
        fee = get_price_at_time("wrxinr", trade["time"])
        trade["fee_in_inr"] = trade["fee"]*fee
        return trade
    elif "inr" in trade["fee_currency"].lower():
        trade["fee_in_inr"] = trade["fee"]
        return trade
    elif "usdt" in trade["fee_currency"].lower():
        fee = get_price_at_time("usdtinr", trade["time"])
        trade["fee_in_inr"] = trade["fee"]*fee
        return trade
    else:
        return "shit"


def process_order(trade):
    if "inr" in trade["market"].lower():
        trade["price_in_inr"] = trade["price"]
        return trade
    elif "usdt" in trade["market"].lower():
        price = get_price_at_time("usdtinr", trade["time"])
        trade["price_in_inr"] = trade["price"]*price
        return trade
    elif "wrx" in trade["market"].lower():
        price = get_price_at_time("wrxinr", trade["time"])
        trade["price_in_inr"] = trade["price"]*price
        return trade
    elif "btc" in trade["market"].lower():
        price = get_price_at_time("btcinr", trade["time"])
        trade["price_in_inr"] = trade["price"]*price
        return trade
    else:
        return "shit"


def add_coin_pk(trade):
    if trade.get('market').endswith('INR'):
        buy_coin, created = Coin.objects.get_or_create(name=trade.get('market').replace("INR", ""))
        sell_coin, created = Coin.objects.get_or_create(name="INR")
    elif trade.get('market').endswith('USDT'):
        buy_coin, created = Coin.objects.get_or_create(name=trade.get('market').replace("USDT", ""))
        sell_coin, created = Coin.objects.get_or_create(name="USDT")
    elif trade.get('market').endswith('WRX'):
        buy_coin, created = Coin.objects.get_or_create(name=trade.get('market').replace("WRX", ""))
        sell_coin, created = Coin.objects.get_or_create(name="WRX")
    elif trade.get('market').endswith('BTC'):
        buy_coin, created = Coin.objects.get_or_create(name=trade.get('market').replace("BTC", ""))
        sell_coin, created = Coin.objects.get_or_create(name="BTC")
    fee_coin, created = Coin.objects.get_or_create(name=trade.get('fee_currency'))
    if trade["type"] == "Buy":
        trade["buy_coin"] = buy_coin.pk
        trade["sell_coin"] = sell_coin.pk
    else:
        trade["buy_coin"] = sell_coin.pk
        trade["sell_coin"] = buy_coin.pk

    trade["fee_coin"] = fee_coin.pk
    return trade


def price_in_inr(orders):
    # pool = Pool(processes=len(orders))
    # orders = pool.map(process_order, orders)
    # results = pool.map(get_network_fee, orders)
    # pool.close()
    # pool.join()
    a = []
    for order in orders:
        temp = process_order(order)
        temp = get_network_fee(temp)
        a.append(temp)
    # orders = process_order(orders)
    # results = get_network_fee(orders)
    results = list(map(add_coin_pk, a))


    # orders = process_order(orders)
    # results = get_network_fee(orders)
    # results = add_coin_pk(results)
    return sorted(results, key=lambda i: i['time'])
